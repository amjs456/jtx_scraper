from playwright.sync_api import sync_playwright, Playwright
import csv
import sys
import openpyxl
from openpyxl.styles import Font
import re


JTX_URL = "https://www.jointex.co.jp/DFiDtl001Servlet?orderCode="

items=[]

def create_order_codes(csv_file_name):
    with open(csv_file_name, newline='') as f:
        spamreader = csv.reader(f)
        order_codes = []
        for row in spamreader:
            order_codes.append(row[0])
        return order_codes


def run(Playwright, order_codes):
    chromium = Playwright.chromium
    browser = chromium.launch()
    page = browser.new_page()
    items=[]
    item_n = len(order_codes)
    for i, order_code in enumerate(order_codes):
        page.goto(JTX_URL+order_code)
        item_code = maker = url = name = price = is_price_red = None

        if not page.title()=="エラー画面":
            price = page.locator("div.rBox aside p.txt02").nth(1).inner_html().replace(" ", "").replace("\n", "").replace("円", "").replace(",", "")
            if price.startswith("<font"):
                price=price.replace("<fontcolor=\"#FF0000\">", "").replace("</font>", "")
                is_price_red = True
            item_code = page.locator("div.info p.txt span").nth(3).inner_html()[7:]
            maker = page.locator("div.info p.txt span").nth(0).inner_html()[6:]
            url = page.url
            name = page.locator("div.info h2").inner_html()

            is_discontinued = False
            has_discontinuation_date = page.locator('tr', has=page.locator('th:has-text("販売中止予定日")')).locator('td').filter(has_text=re.compile(r'^\d{4}/\d{2}/\d{2}$')).count() > 0
            has_discontinued_date = page.locator('tr', has=page.locator('th:has-text("販売中止日")')).locator('td').filter(has_text=re.compile(r'^\d{4}/\d{2}/\d{2}$')).count() > 0
            is_discontinued_img_exist =  page.locator('img[src="img/icon_img/ricon_13.png"], img[src="img/icon_img/ricon_14.png"], img[src="img/icon_img/ricon_20.png"]').count() > 0
            if has_discontinuation_date or has_discontinued_date or is_discontinued_img_exist:
                is_discontinued = True

        item_info = {
            "order_code":order_code,
            "maker":maker,
            "name":name,
            "price":price,
            "is_price_red":is_price_red,
            "item_code":item_code,
            "is_discontinued":is_discontinued,
            "url":url
        }
        items.append(item_info)
        header_list = list(item_info.keys())
        print(str(i+1)+"/"+str(item_n)+" completed!")
    browser.close()
    return items

def save_excel_file(items, filename):
    red_price_letter = "#"
    header = ["注文コード", "クラウンかも", "キスパかも", "オーダーかも", "それ以外", "メーカー", "商品名", "品番", "価格", "販売中止", "URL"]
    data = [header]
    price_idx = header.index("価格")

    for item in items:
        record = []
        record.append(item["order_code"])
        if not len(item["order_code"])==6:
            if len(item["order_code"])==5:
                maybe = ["1", "0", "0", "0"]
            elif len(item["order_code"])==8:
                maybe = ["0", "1", "0", "0"]
            elif len(item["order_code"])==9:
                maybe = ["0", "0", "1", "0"]
            else:
                maybe = ["0", "0", "0", "1"]
            record = record + maybe
        else:
            record = record + ["0", "0", "0", "0"]
        record.append(item["maker"])
        record.append(item["name"])
        record.append(item["item_code"])
        if item["is_price_red"]:
            record.append(red_price_letter+str(item["price"]))
        else:
            record.append(item["price"])
        if item["is_discontinued"]:
            record.append("1")
        else:
            record.append("0")
        record.append(item["url"])
        data.append(record)

    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws.title = "JOINTEX"
    for i, rec in enumerate(data):
        for j, col in enumerate(rec):
            col_letter = chr(ord("A")+j)
            address = f"{col_letter}{i+1}"
            if j==price_idx:
                if col:
                    if col.startswith(red_price_letter):
                        ws[address] = col.replace(red_price_letter, "")
                        ws[address].font = Font(color="ff0000")
                    else:
                        ws[address] = col
            else:
                ws[address] = col
    wb.save(filename=filename)


with sync_playwright() as playwright:
    csv_file_name = sys.argv[1]
    if not csv_file_name.endswith(".csv"):
        csv_file_name = csv_file_name + ".csv"
    order_codes=  create_order_codes(csv_file_name)
    items = run(playwright, order_codes)
    save_excel_file(items, "sample.xlsx")
